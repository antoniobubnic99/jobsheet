"""The workbook writer.

This is the strictest test file in the project, because it guards the one thing
the user cannot get back if it breaks: their own notes, ticks and statuses.

The predecessor lost 88 hand-placed ticks in a single run. `test_ticks_survive_*`
and `test_verification_failure_restores_the_backup` exist so that cannot happen
twice quietly.
"""

from __future__ import annotations

import random
import zipfile
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from jobsheet.core.models import ApplicationStatus, Posting
from jobsheet.sheet.checkbox import FPB_PART, CheckboxError, apply_native_checkboxes
from jobsheet.sheet.layout import (
    ColumnKind,
    ColumnSpec,
    ConditionalRule,
    SheetLayout,
    classic_checkboxes_layout,
    default_layout,
    minimal_layout,
)
from jobsheet.sheet.row import JobRow
from jobsheet.sheet.writer import (
    MAX_ROWS,
    SheetLockedError,
    VerificationFailedError,
    create_empty,
    load,
    prune_backups,
    read_layout,
    save,
)

FOUND = date(2026, 8, 24)


def job(number: int, **overrides: Any) -> JobRow:
    data: dict[str, Any] = {
        "posting": Posting(
            source_id="rss",
            title=f"Job {number}",
            url=f"https://example.test/j/{number}",
            company=f"Company {number}",
            location="Zagreb",
            posted_at=date(2026, 8, (number % 28) + 1),
        ),
        "found_at": FOUND,
        "category": "GIS",
        "note": "found via feed",
    }
    return JobRow(**(data | overrides))


@pytest.fixture
def book(tmp_path: Path) -> Path:
    return tmp_path / "jobs.xlsx"


# ----------------------------------------------------------------- round trip


class TestRoundTrip:
    def test_empty_workbook_has_a_header_and_no_rows(self, book: Path) -> None:
        layout = create_empty(book)
        rows, _ = load(book)
        assert rows == []
        assert read_layout(book) == layout

    def test_rows_survive_a_write_and_read(self, book: Path) -> None:
        create_empty(book)
        save(book, [job(n) for n in range(1, 26)])
        rows, _ = load(book)
        assert len(rows) == 25
        assert {r.posting.title for r in rows} == {f"Job {n}" for n in range(1, 26)}

    def test_dates_come_back_as_dates_not_strings(self, book: Path) -> None:
        """Real dates keep Excel's sorting and filtering working."""
        create_empty(book)
        with_deadline = job(1)
        with_deadline = job(
            1, posting=with_deadline.posting.model_copy(update={"deadline": date(2026, 9, 12)})
        )
        save(book, [with_deadline])
        rows, _ = load(book)
        assert rows[0].found_at == FOUND
        assert rows[0].posting.deadline == date(2026, 9, 12)

    def test_link_is_a_real_hyperlink(self, book: Path) -> None:
        create_empty(book)
        save(book, [job(1)])
        workbook = openpyxl.load_workbook(book)
        url_column = default_layout().index_of("url")
        assert url_column is not None
        assert workbook["Jobs"].cell(2, url_column).hyperlink is not None
        workbook.close()

    def test_typed_link_label_survives(self, book: Path) -> None:
        """Some people rename a link to something they recognise."""
        create_empty(book)
        save(book, [job(1, link_text="the good one at Company 1")])
        rows, _ = load(book)
        assert rows[0].link_text == "the good one at Company 1"
        assert rows[0].posting.url == "https://example.test/j/1"

    def test_layout_is_embedded_in_the_file(self, book: Path) -> None:
        layout = minimal_layout()
        create_empty(book, layout)
        assert read_layout(book) == layout

    def test_empty_rows_are_skipped(self, book: Path) -> None:
        create_empty(book)
        save(book, [job(1)])
        workbook = openpyxl.load_workbook(book)
        workbook["Jobs"].cell(10, 2, "   ")
        workbook.save(book)
        workbook.close()
        rows, _ = load(book)
        assert len(rows) == 1


# ------------------------------------------------------------------ integrity


class TestUserDataIntegrity:
    def test_statuses_survive_a_rewrite(self, book: Path) -> None:
        create_empty(book)
        rows = [job(n) for n in range(1, 11)]
        rows[0] = job(1, status=ApplicationStatus.APPLIED)
        rows[3] = job(4, status=ApplicationStatus.INTERVIEW)
        rows[7] = job(8, status=ApplicationStatus.REJECTED)
        save(book, rows)

        back, _ = load(book)
        by_title = {r.posting.title: r.status for r in back}
        assert by_title["Job 1"] is ApplicationStatus.APPLIED
        assert by_title["Job 4"] is ApplicationStatus.INTERVIEW
        assert by_title["Job 8"] is ApplicationStatus.REJECTED

    def test_ticks_survive_a_rewrite(self, book: Path) -> None:
        """The regression test for the incident that started all of this."""
        layout = classic_checkboxes_layout()
        create_empty(book, layout)
        rows = [job(n) for n in range(1, 21)]
        rows[0] = job(1, user_values={"applied": True, "interview": True})
        rows[5] = job(6, user_values={"applied": True})
        rows[9] = job(10, user_values={"not_applied": True})
        save(book, rows, layout)

        back, _ = load(book, layout)
        ticks = {r.posting.title: {k for k, v in r.user_values.items() if v is True} for r in back}
        assert ticks["Job 1"] == {"applied", "interview"}
        assert ticks["Job 6"] == {"applied"}
        assert ticks["Job 10"] == {"not_applied"}

    def test_ticks_stay_with_their_row_across_a_re_sort(self, book: Path) -> None:
        """Sorting is where the original bug lived: values slid between rows.

        Rows are written in a different order than they arrive, so if ticks were
        written column-wise they would attach to whichever job happened to land
        in that position.
        """
        layout = classic_checkboxes_layout()
        create_empty(book, layout)
        rows = [job(n) for n in range(1, 31)]
        marked = {3, 11, 17, 29}
        for index in marked:
            rows[index - 1] = job(index, user_values={"applied": True})
        save(book, rows, layout)

        back, _ = load(book, layout)
        applied = {r.posting.title for r in back if r.user_values.get("applied") is True}
        assert applied == {f"Job {n}" for n in marked}

    def test_randomised_user_edits_are_never_lost(self, book: Path) -> None:
        """Property-style check over many shapes of user data at once."""
        rng = random.Random(20260824)
        layout = default_layout()
        create_empty(book, layout)

        rows = []
        expected: dict[str, tuple[str, str]] = {}
        for n in range(1, 121):
            status = rng.choice(list(ApplicationStatus))
            note = rng.choice(["", "call them", "waiting", "sent CV", "ask about salary"])
            rows.append(job(n, status=status, user_values={"my_notes": note}))
            expected[f"Job {n}"] = (str(status), note)
        save(book, rows, layout)

        back, _ = load(book, layout)
        assert len(back) == 120
        actual = {
            r.posting.title: (str(r.status), r.user_values.get("my_notes") or "") for r in back
        }
        assert actual == expected

    def test_repeated_saves_are_stable(self, book: Path) -> None:
        """Five runs in a row must not erode anything."""
        layout = default_layout()
        create_empty(book, layout)
        rows = [job(n, status=ApplicationStatus.APPLIED) for n in range(1, 16)]
        save(book, rows, layout)

        for _ in range(5):
            current, current_layout = load(book, layout)
            save(book, current, current_layout)

        final, _ = load(book, layout)
        assert len(final) == 15
        assert all(r.status is ApplicationStatus.APPLIED for r in final)


# --------------------------------------------------------- the user's own edits


class TestEditsMadeInExcel:
    def test_a_column_added_by_hand_in_excel_is_kept(self, book: Path) -> None:
        """Typing a header into Excel is the user asking to track something."""
        layout = minimal_layout()
        create_empty(book, layout)
        save(book, [job(1), job(2)], layout)

        workbook = openpyxl.load_workbook(book)
        sheet = workbook["Jobs"]
        extra = len(layout.columns) + 1
        sheet.cell(1, extra, "Salary guess")
        sheet.cell(2, extra, "2200 EUR")
        workbook.save(book)
        workbook.close()

        rows, discovered_layout = load(book, layout)
        keys = [c.key for c in discovered_layout.columns]
        assert "custom:salary guess" in keys
        assert any(r.user_values.get("custom:salary guess") == "2200 EUR" for r in rows)

    def test_hand_added_column_survives_the_next_save(self, book: Path) -> None:
        layout = minimal_layout()
        create_empty(book, layout)
        save(book, [job(1)], layout)

        workbook = openpyxl.load_workbook(book)
        sheet = workbook["Jobs"]
        extra = len(layout.columns) + 1
        sheet.cell(1, extra, "Salary guess")
        sheet.cell(2, extra, "2200 EUR")
        workbook.save(book)
        workbook.close()

        rows, grown = load(book, layout)
        save(book, rows, grown)

        after, _ = load(book, grown)
        assert after[0].user_values.get("custom:salary guess") == "2200 EUR"

    def test_reordered_columns_are_matched_by_header(self, book: Path) -> None:
        """Dragging a column in Excel must not scramble the next read."""
        layout = minimal_layout()
        create_empty(book, layout)
        save(book, [job(1)], layout)

        reordered = SheetLayout(
            sheet_name=layout.sheet_name,
            columns=[layout.columns[2], layout.columns[0], layout.columns[1], layout.columns[3]],
        )
        rows, _ = load(book, reordered)
        assert rows[0].posting.title == "Job 1"
        assert rows[0].posting.company == "Company 1"

    def test_status_typed_in_excel_is_read_back(self, book: Path) -> None:
        layout = default_layout()
        create_empty(book, layout)
        save(book, [job(1)], layout)

        workbook = openpyxl.load_workbook(book)
        column = layout.index_of("status")
        assert column is not None
        workbook["Jobs"].cell(2, column, "offer")
        workbook.save(book)
        workbook.close()

        rows, _ = load(book, layout)
        assert rows[0].status is ApplicationStatus.OFFER

    def test_unknown_status_word_is_tolerated(self, book: Path) -> None:
        """A typo in the status cell must not crash the run."""
        layout = default_layout()
        create_empty(book, layout)
        save(book, [job(1)], layout)

        workbook = openpyxl.load_workbook(book)
        column = layout.index_of("status")
        assert column is not None
        workbook["Jobs"].cell(2, column, "maybe???")
        workbook.save(book)
        workbook.close()

        rows, _ = load(book, layout)
        assert rows[0].status is ApplicationStatus.NEW


# -------------------------------------------------------------- refusing to write


class TestRefusalAndRollback:
    def test_refuses_while_the_workbook_is_open(self, book: Path) -> None:
        create_empty(book)
        save(book, [job(1)])
        book.with_name(f"~${book.name}").write_text("")
        with pytest.raises(SheetLockedError):
            save(book, [job(1), job(2)])

    def test_refuses_an_absurd_number_of_rows(self, book: Path) -> None:
        create_empty(book)
        with pytest.raises(ValueError, match="refusing to write"):
            save(book, [job(1)] * (MAX_ROWS + 1))

    def test_verification_failure_restores_the_backup(
        self, book: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """If the file that comes back is wrong, the old one must come back."""
        layout = default_layout()
        create_empty(book, layout)
        good = [job(n, status=ApplicationStatus.APPLIED) for n in range(1, 6)]
        save(book, good, layout)
        before = book.read_bytes()

        import jobsheet.sheet.writer as writer

        def broken(path: Path | str, given: SheetLayout | None = None) -> Any:
            return [], given or layout

        monkeypatch.setattr(writer, "load", broken)
        with pytest.raises(VerificationFailedError):
            save(book, good, layout)
        monkeypatch.undo()

        assert book.read_bytes() == before
        rows, _ = load(book, layout)
        assert len(rows) == 5
        assert all(r.status is ApplicationStatus.APPLIED for r in rows)


# ------------------------------------------------------------------- formatting


class TestFormatting:
    def test_conditional_formatting_covers_every_row(self, book: Path) -> None:
        """The predecessor pinned this to A2:I500 while writing 5000 rows."""
        layout = classic_checkboxes_layout()
        create_empty(book, layout)
        save(book, [job(n) for n in range(1, 601)], layout)

        workbook = openpyxl.load_workbook(book)
        ranges = [str(r.sqref) for r in workbook["Jobs"].conditional_formatting]
        workbook.close()
        assert ranges
        assert all(r.endswith("601") for r in ranges), ranges

    def test_autofilter_and_freeze_follow_the_layout(self, book: Path) -> None:
        layout = minimal_layout()
        create_empty(book, layout)
        save(book, [job(n) for n in range(1, 4)], layout)

        workbook = openpyxl.load_workbook(book)
        sheet = workbook["Jobs"]
        assert sheet.auto_filter.ref == "A1:D4"
        assert sheet.freeze_panes == "A2"
        workbook.close()

    def test_freeze_and_filter_can_be_turned_off(self, book: Path) -> None:
        layout = minimal_layout().model_copy(update={"freeze_header": False, "autofilter": False})
        create_empty(book, layout)
        save(book, [job(1)], layout)

        workbook = openpyxl.load_workbook(book)
        sheet = workbook["Jobs"]
        assert sheet.freeze_panes is None
        assert sheet.auto_filter.ref is None
        workbook.close()

    def test_custom_colours_reach_the_file(self, book: Path) -> None:
        layout = SheetLayout(
            columns=[
                ColumnSpec(key="title", label="Position"),
                ColumnSpec(key="status", label="Status", kind=ColumnKind.STATUS),
            ],
            rules=[ConditionalRule(when_column="status", equals="offer", fill="FFE699")],
        )
        create_empty(book, layout)
        save(book, [job(1)], layout)

        workbook = openpyxl.load_workbook(book)
        rules = list(workbook["Jobs"].conditional_formatting)
        colours = [r.rules[0].dxf.fill.bgColor.rgb for r in rules]
        workbook.close()
        assert "FFFFE699" in colours

    def test_column_widths_come_from_the_layout(self, book: Path) -> None:
        layout = minimal_layout()
        create_empty(book, layout)
        workbook = openpyxl.load_workbook(book)
        assert workbook["Jobs"].column_dimensions["A"].width == layout.columns[0].width
        workbook.close()


# --------------------------------------------------------------- checkboxes


class TestNativeCheckboxes:
    def test_the_feature_part_is_written(self, book: Path) -> None:
        layout = classic_checkboxes_layout()
        create_empty(book, layout)
        save(book, [job(1, user_values={"applied": True})], layout)
        with zipfile.ZipFile(book) as archive:
            assert FPB_PART in archive.namelist()

    def test_content_type_and_relationship_are_registered(self, book: Path) -> None:
        """Miss either and Excel calls the file corrupt."""
        layout = classic_checkboxes_layout()
        create_empty(book, layout)
        save(book, [job(1)], layout)
        with zipfile.ZipFile(book) as archive:
            content_types = archive.read("[Content_Types].xml").decode()
            rels = archive.read("xl/_rels/workbook.xml.rels").decode()
        assert "featurepropertybag" in content_types
        assert "FeaturePropertyBag" in rels

    def test_workbook_still_opens_afterwards(self, book: Path) -> None:
        """The zip surgery must leave a file openpyxl can still read."""
        layout = classic_checkboxes_layout()
        create_empty(book, layout)
        save(book, [job(n) for n in range(1, 5)], layout)
        workbook = openpyxl.load_workbook(book)
        assert workbook["Jobs"].max_row == 5
        workbook.close()

    def test_no_checkbox_columns_means_no_surgery(self, book: Path) -> None:
        layout = minimal_layout()
        create_empty(book, layout)
        save(book, [job(1)], layout)
        with zipfile.ZipFile(book) as archive:
            assert FPB_PART not in archive.namelist()

    def test_missing_sheet_name_is_reported(self, book: Path) -> None:
        create_empty(book, minimal_layout())
        with pytest.raises(CheckboxError, match="no sheet named"):
            apply_native_checkboxes(book, column_indices=[1], sheet_name="Nope")

    def test_no_columns_is_a_no_op(self, book: Path) -> None:
        create_empty(book, minimal_layout())
        assert apply_native_checkboxes(book, column_indices=[]) == -1


# ---------------------------------------------------------------------- backups


class TestBackups:
    def test_a_backup_is_made_before_overwriting(self, book: Path) -> None:
        create_empty(book)
        save(book, [job(1)])
        report = save(book, [job(1), job(2)])
        assert report.backup is not None
        assert report.backup.exists()

    def test_first_write_has_nothing_to_back_up(self, book: Path) -> None:
        create_empty(book)
        book.unlink()
        assert save(book, [job(1)]).backup is None

    def test_backups_are_pruned(self, book: Path, tmp_path: Path) -> None:
        """The predecessor never pruned and accumulated two dozen copies."""
        backups = tmp_path / "backups"
        create_empty(book)
        for n in range(8):
            save(book, [job(n + 1)], backup_dir=backups, keep_backups=3)
        assert len(list(backups.glob("*.xlsx"))) <= 3

    def test_pruning_a_missing_directory_is_harmless(self, tmp_path: Path) -> None:
        assert prune_backups(tmp_path / "jobs.xlsx", tmp_path / "nope") == 0
