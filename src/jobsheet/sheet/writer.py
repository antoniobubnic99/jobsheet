"""Reading and writing the workbook, without ever losing the user's own work.

The safety envelope around every save is the reason this module exists, and it is
not theoretical. In the predecessor project a single run destroyed 88 hand-placed
ticks: re-sorting rewrote the sheet column by column, so values slid up and down
independently of the rows they belonged to. Nothing raised, nothing looked wrong,
and the damage was noticed days later.

Two rules prevent a repeat, and both are structural rather than careful:

* Rows are sorted and written as whole objects. Cells are never moved.
* Every save is: refuse if the file is open -> count the user's values -> back up
  -> write -> read back -> compare the counts -> restore the backup if they
  differ. A save that cannot prove it preserved the user's work is not a save.

The layout is embedded in the file itself, so a workbook the user has rearranged
by hand still reads correctly on the next run -- and a column they added directly
in Excel is picked up and preserved rather than wiped.
"""

from __future__ import annotations

import json
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.packaging.custom import StringProperty
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from jobsheet.core.dates import parse_date
from jobsheet.core.models import ApplicationStatus, Posting
from jobsheet.sheet.checkbox import apply_native_checkboxes
from jobsheet.sheet.layout import ColumnKind, ColumnSpec, SheetLayout, default_layout
from jobsheet.sheet.row import JobRow, cell_value, sort_key
from jobsheet.sheet.theme import resolve_theme

__all__ = [
    "LAYOUT_PROPERTY",
    "MAX_ROWS",
    "SaveReport",
    "SheetLockedError",
    "VerificationFailedError",
    "create_empty",
    "export_layout",
    "is_locked",
    "load",
    "make_backup",
    "prune_backups",
    "read_layout",
    "save",
]

LAYOUT_PROPERTY = "JobSheetLayout"

# A guard against a bug that fills the sheet to Excel's limit. Anything past this
# is a runaway loop, not a job search.
MAX_ROWS = 50_000

DATE_FORMAT = "yyyy-mm-dd"

_STATUSES = [str(s) for s in ApplicationStatus]


class SheetLockedError(RuntimeError):
    """The workbook is open in Excel, which holds an exclusive lock."""


class VerificationFailedError(RuntimeError):
    """The file written back did not match what went in. The backup was restored."""


class SaveReport:
    """What a save actually did, for the caller to report to the user."""

    def __init__(self, rows: int, user_values: int, backup: Path | None) -> None:
        self.rows = rows
        self.user_values = user_values
        self.backup = backup

    def __repr__(self) -> str:
        return f"SaveReport(rows={self.rows}, user_values={self.user_values})"


# ------------------------------------------------------------------ lock check


def _lock_path(path: Path) -> Path:
    """Excel's owner file: `~$name.xlsx` beside the workbook while it is open."""
    return path.with_name(f"~${path.name}")


def is_locked(path: Path | str) -> bool:
    return _lock_path(Path(path)).exists()


# ---------------------------------------------------------------- layout in file


def _embed_layout(workbook: openpyxl.Workbook, layout: SheetLayout) -> None:
    """Store the layout inside the workbook so the next run can read it back."""
    properties = workbook.custom_doc_props
    # `CustomPropertyList` has no removal API, so replace the backing list.
    properties.props = [p for p in properties.props if p.name != LAYOUT_PROPERTY]
    properties.append(StringProperty(name=LAYOUT_PROPERTY, value=layout.model_dump_json()))


def read_layout(path: Path | str) -> SheetLayout | None:
    """The layout a workbook was written with, if it carries one."""
    workbook = openpyxl.load_workbook(Path(path))
    try:
        properties = workbook.custom_doc_props
        if LAYOUT_PROPERTY not in properties.names:
            return None
        stored = properties[LAYOUT_PROPERTY]
        if not stored.value:
            return None
        return SheetLayout.model_validate_json(str(stored.value))
    except (ValueError, TypeError, KeyError):
        # A corrupted or hand-edited property must not stop the user opening
        # their own file; header matching takes over from here.
        return None
    finally:
        workbook.close()


# ------------------------------------------------------------------- reading


def _header_map(sheet: Worksheet) -> dict[str, int]:
    """Header label (case-folded) -> 1-based column index, as the file has it."""
    found: dict[str, int] = {}
    for index, cell in enumerate(next(sheet.iter_rows(max_row=1), ()), start=1):
        label = str(cell.value or "").strip().casefold()
        if label and label not in found:
            found[label] = index
    return found


def _resolve_columns(
    sheet: Worksheet, layout: SheetLayout
) -> tuple[dict[str, int], list[ColumnSpec]]:
    """Work out where each column actually sits in this particular file.

    Matching is by header label first, because that survives the user dragging
    columns around in Excel. Columns whose label was also renamed fall back to
    their position in the layout.

    Any header the layout does not know about becomes a custom column. That is
    deliberate: a column typed straight into Excel is the user telling us they
    want to track something, and the correct response is to keep it.
    """
    headers = _header_map(sheet)
    resolved: dict[str, int] = {}
    claimed: set[int] = set()

    for column in layout.columns:
        index = headers.get(column.label.casefold())
        if index is not None and index not in claimed:
            resolved[column.key] = index
            claimed.add(index)

    for position, column in enumerate(layout.columns, start=1):
        if column.key in resolved:
            continue
        if position not in claimed:
            resolved[column.key] = position
            claimed.add(position)

    discovered: list[ColumnSpec] = []
    for label, index in headers.items():
        if index in claimed:
            continue
        original = str(sheet.cell(1, index).value or "").strip()
        key = f"custom:{label}"
        discovered.append(ColumnSpec(key=key, label=original or label, kind=ColumnKind.TEXT))
        resolved[key] = index
        claimed.add(index)

    return resolved, discovered


def _read_cell(sheet: Worksheet, row: int, index: int, kind: ColumnKind) -> Any:
    cell = sheet.cell(row, index)
    value = cell.value

    if kind is ColumnKind.URL:
        # The visible text may be a label the user typed; the hyperlink target is
        # the real address. Prefer the target, keep the label separately.
        return cell.hyperlink.target if cell.hyperlink else value
    if kind is ColumnKind.CHECKBOX:
        # Strict identity: the string "TRUE" is not a tick, it is a typo.
        return value is True
    if kind is ColumnKind.DATE:
        if isinstance(value, datetime):
            return value.date()
        return parse_date(value)
    return value


def load(path: Path | str, layout: SheetLayout | None = None) -> tuple[list[JobRow], SheetLayout]:
    """Read every row back, together with the layout the file is really using."""
    path = Path(path)
    layout = layout or read_layout(path) or default_layout()

    workbook = openpyxl.load_workbook(path)
    try:
        sheet = (
            workbook[layout.sheet_name]
            if layout.sheet_name in workbook.sheetnames
            else workbook.active
        )
        if sheet is None:
            return [], layout

        columns, discovered = _resolve_columns(sheet, layout)
        if discovered:
            layout = layout.model_copy(update={"columns": [*layout.columns, *discovered]})

        by_key = {c.key: c for c in layout.columns}
        rows: list[JobRow] = []

        for number in range(2, min(sheet.max_row, MAX_ROWS) + 1):
            values = {
                key: _read_cell(sheet, number, index, by_key[key].kind)
                for key, index in columns.items()
                if key in by_key
            }
            # A row with no title and no link is an empty row, not a job.
            if not values.get("title") and not values.get("url"):
                continue

            user_values = {
                key: values.get(key)
                for key, column in by_key.items()
                if column.user_owned and key != "status"
            }
            link_index = columns.get("url")
            link_text = str(sheet.cell(number, link_index).value or "") if link_index else ""

            rows.append(
                JobRow(
                    posting=Posting(
                        source_id=str(values.get("source") or "imported"),
                        title=str(values.get("title") or ""),
                        url=str(values.get("url") or ""),
                        company=str(values.get("company") or ""),
                        location=str(values.get("location") or ""),
                        region=str(values.get("region") or ""),
                        employment_type=str(values.get("employment_type") or ""),
                        education=str(values.get("education") or ""),
                        salary=str(values.get("salary") or ""),
                        posted_at=values.get("posted_at"),
                        deadline=values.get("deadline"),
                    ),
                    found_at=values.get("found_at") or date.today(),
                    category=str(values.get("category") or ""),
                    note=str(values.get("note") or ""),
                    status=_read_status(values.get("status")),
                    user_values=user_values,
                    link_text=link_text,
                )
            )
        return rows, layout
    finally:
        workbook.close()


def _read_status(value: Any) -> ApplicationStatus:
    """Tolerate anything in the status cell; an unknown word means untouched."""
    try:
        return ApplicationStatus(str(value or "").strip().lower())
    except ValueError:
        return ApplicationStatus.NEW


# ------------------------------------------------------------------- writing


def _write_sheet(sheet: Worksheet, rows: list[JobRow], layout: SheetLayout) -> None:
    """Lay out the header and every row. Nothing here knows a column count."""
    theme = resolve_theme(layout.theme)

    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row)

    header_fill = PatternFill("solid", fgColor=theme.header_fill)
    header_font = Font(bold=True, color=theme.header_text if theme.is_dark_header else "FF111111")
    link_font = Font(color=theme.link_text, underline="single")
    zebra_fill = PatternFill("solid", fgColor=theme.zebra_fill)

    for index, column in enumerate(layout.columns, start=1):
        cell = sheet.cell(1, index, column.label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = column.width

    # Sorted as whole objects. This single line is what keeps a tick attached to
    # its job; sorting cells instead is what destroyed the predecessor's data.
    ordered = sorted(
        rows,
        key=lambda r: sort_key(r, layout.sort_by),
        reverse=layout.sort_descending,
    )

    for offset, row in enumerate(ordered, start=2):
        for index, column in enumerate(layout.columns, start=1):
            cell = sheet.cell(offset, index)
            value = cell_value(row, column.key)

            if column.kind is ColumnKind.URL and value:
                cell.value = row.link_text or value
                cell.hyperlink = str(value)
                cell.font = link_font
            elif column.kind is ColumnKind.CHECKBOX:
                cell.value = bool(value)
            elif column.kind is ColumnKind.DATE:
                cell.value = value
                if value is not None:
                    cell.number_format = DATE_FORMAT
            else:
                cell.value = value

            if column.wrap:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if layout.zebra and offset % 2 == 0 and column.kind is not ColumnKind.CHECKBOX:
                cell.fill = zebra_fill

    last_row = max(len(ordered) + 1, 2)
    last_column = get_column_letter(len(layout.columns))

    if layout.freeze_header:
        sheet.freeze_panes = "A2"
    if layout.autofilter:
        sheet.auto_filter.ref = f"A1:{last_column}{last_row}"

    _apply_status_validation(sheet, layout, last_row)
    _apply_rules(sheet, layout, last_row, last_column)


def _apply_status_validation(sheet: Worksheet, layout: SheetLayout, last_row: int) -> None:
    """A dropdown, so the status column stays a vocabulary rather than free text."""
    for index, column in enumerate(layout.columns, start=1):
        if column.kind is not ColumnKind.STATUS:
            continue
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(_STATUSES)}"',
            allow_blank=True,
        )
        sheet.add_data_validation(validation)
        letter = get_column_letter(index)
        validation.add(f"{letter}2:{letter}{max(last_row, 2)}")


def _apply_rules(sheet: Worksheet, layout: SheetLayout, last_row: int, last_column: str) -> None:
    """Rebuild conditional formatting from the layout, over the real used range.

    The predecessor pinned this to `A2:I500` while allowing 5000 rows, so row
    501 onwards silently lost its colours. Deriving the range removes the class
    of bug rather than moving the number.
    """
    sheet.conditional_formatting = ConditionalFormattingList()
    if not layout.rules or last_row < 2:
        return

    span = f"A2:{last_column}{last_row}"
    for rule in layout.rules:
        index = layout.index_of(rule.when_column)
        if index is None:
            continue
        letter = get_column_letter(index)
        if isinstance(rule.equals, bool):
            formula = f"${letter}2={'TRUE' if rule.equals else 'FALSE'}"
        else:
            formula = f'${letter}2="{rule.equals}"'
        sheet.conditional_formatting.add(
            span,
            FormulaRule(
                formula=[formula],
                stopIfTrue=rule.stop_if_true,
                fill=PatternFill(start_color=rule.argb, end_color=rule.argb, fill_type="solid"),
            ),
        )


# -------------------------------------------------------------------- backups


def make_backup(path: Path, backup_dir: Path, label: str = "") -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    suffix = f"-{label}" if label else ""
    destination = backup_dir / f"{path.stem}-backup-{stamp}{suffix}{path.suffix}"
    shutil.copy2(path, destination)
    return destination


def prune_backups(path: Path, backup_dir: Path, keep: int = 20) -> int:
    """Keep the most recent backups and delete the rest.

    The predecessor never pruned and accumulated two dozen copies of the same
    workbook, several written within the same minute.
    """
    if not backup_dir.exists():
        return 0
    existing = sorted(
        backup_dir.glob(f"{path.stem}-backup-*{path.suffix}"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    removed = 0
    for stale in existing[keep:]:
        stale.unlink(missing_ok=True)
        removed += 1
    return removed


# ---------------------------------------------------------------------- save


def create_empty(path: Path | str, layout: SheetLayout | None = None) -> SheetLayout:
    """Make a workbook with a header and no rows."""
    path = Path(path)
    layout = layout or default_layout()
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = layout.sheet_name
    _write_sheet(sheet, [], layout)
    _embed_layout(workbook, layout)
    workbook.save(path)
    workbook.close()
    return layout


def save(
    path: Path | str,
    rows: list[JobRow],
    layout: SheetLayout | None = None,
    *,
    backup_dir: Path | str | None = None,
    label: str = "",
    keep_backups: int = 20,
) -> SaveReport:
    """Write every row, and prove afterwards that nothing of the user's was lost.

    Raises `SheetLockedError` if the workbook is open, and `VerificationFailedError` if the
    file that came back does not match what went in -- in which case the backup
    has already been restored and the workbook on disk is untouched.
    """
    path = Path(path)
    layout = layout or (read_layout(path) if path.exists() else None) or default_layout()
    backups = Path(backup_dir) if backup_dir else path.parent / "backups"

    if is_locked(path):
        raise SheetLockedError(
            f"{path.name} is open in Excel. Close it and run again -- "
            "writing now would either fail or lose your edits."
        )
    if len(rows) > MAX_ROWS:
        raise ValueError(f"refusing to write {len(rows)} rows; the limit is {MAX_ROWS}")

    expected_rows = len(rows)
    expected_values = sum(row.user_value_count() for row in rows)
    backup = make_backup(path, backups, label) if path.exists() else None

    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = layout.sheet_name
        _write_sheet(sheet, rows, layout)
        _embed_layout(workbook, layout)
        workbook.save(path)
        workbook.close()

        checkbox_indices = [
            index
            for index, column in enumerate(layout.columns, start=1)
            if column.kind is ColumnKind.CHECKBOX
        ]
        if checkbox_indices:
            apply_native_checkboxes(
                path, column_indices=checkbox_indices, sheet_name=layout.sheet_name
            )

        written, _ = load(path, layout)
        actual_values = sum(row.user_value_count() for row in written)
        if len(written) != expected_rows or actual_values != expected_values:
            raise VerificationFailedError(
                f"rows {len(written)}/{expected_rows}, "
                f"your values {actual_values}/{expected_values}"
            )
    except Exception:
        if backup is not None:
            shutil.copy2(backup, path)
        raise

    prune_backups(path, backups, keep=keep_backups)
    return SaveReport(rows=expected_rows, user_values=expected_values, backup=backup)


def export_layout(layout: SheetLayout) -> str:
    """The layout as shareable JSON, for the "save this design" button."""
    return json.dumps(layout.model_dump(mode="json"), indent=2, ensure_ascii=False)
