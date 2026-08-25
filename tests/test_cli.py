"""The `jobsheet` command.

`run` and `export` are what a scheduled task calls at three in the morning with
nobody watching, so the exit codes matter as much as the output: a locked
workbook has to fail loudly rather than appear to have worked.
"""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path
from typing import Any, ClassVar

import openpyxl
import pytest

from jobsheet.cli import build_parser, free_port, main, parse_params
from jobsheet.core.models import ApplicationStatus, Posting
from jobsheet.sheet.layout import ColumnSpec, SheetLayout, default_layout
from jobsheet.sheet.row import JobRow
from jobsheet.sources import registry
from jobsheet.sources.base import FetchContext, Source, SourceManifest
from jobsheet.store.db import Database
from jobsheet.store.tracker import Tracker

FOUND = date(2026, 8, 24)


class CliSource(Source):
    manifest = SourceManifest(id="clifake", name="CLI Fake", homepage="https://example.test")
    postings: ClassVar[list[Posting]] = []
    seen_params: ClassVar[dict[str, Any]] = {}

    async def fetch(self, params: dict[str, Any], ctx: FetchContext) -> list[Posting]:
        type(self).seen_params = dict(params)
        return list(type(self).postings)


@pytest.fixture(autouse=True)
def _register() -> None:
    registry.register(CliSource)
    CliSource.postings = []
    CliSource.seen_params = {}


def ad(number: int, **overrides: Any) -> Posting:
    data: dict[str, Any] = {
        "source_id": "clifake",
        "title": f"GIS Engineer {number}",
        "url": f"https://example.test/j/{number}",
        "company": f"Company {number}",
        "location": "Zagreb",
        "posted_at": FOUND,
    }
    return Posting(**(data | overrides))


def job(number: int, **overrides: Any) -> JobRow:
    data: dict[str, Any] = {"posting": ad(number), "found_at": FOUND, "category": "GIS"}
    return JobRow(**(data | overrides))


@pytest.fixture
def home(tmp_path: Path) -> Path:
    return tmp_path / "home"


def run_cli(home: Path, *argv: str) -> int:
    return main(["--home", str(home), *argv])


def stored(home: Path) -> list[JobRow]:
    with Database(home / "jobsheet.sqlite3") as db:
        return db.all_rows()


# ------------------------------------------------------------------- parsing


class TestParser:
    def test_no_command_means_open_the_interface(self) -> None:
        """A double-clicked launcher passes no arguments at all."""
        assert build_parser().parse_args([]).command is None

    def test_global_options_come_before_the_command(self, tmp_path: Path) -> None:
        args = build_parser().parse_args(["--home", str(tmp_path), "sources"])
        assert args.home == tmp_path
        assert args.command == "sources"

    def test_an_unknown_export_format_is_refused(self) -> None:
        with pytest.raises(SystemExit):
            build_parser().parse_args(["export", "pdf"])


class TestParseParams:
    def test_one_parameter_for_one_source(self) -> None:
        assert parse_params(["hzz:county=8"]) == {"hzz": {"county": "8"}}

    def test_several_parameters_group_by_source(self) -> None:
        assert parse_params(["hzz:county=8", "hzz:page=2", "rss:url=x"]) == {
            "hzz": {"county": "8", "page": "2"},
            "rss": {"url": "x"},
        }

    def test_a_value_may_contain_an_equals_sign(self) -> None:
        """Feed URLs are full of them, so only the first one separates."""
        assert parse_params(["rss:url=https://x.test/f?a=1&b=2"]) == {
            "rss": {"url": "https://x.test/f?a=1&b=2"}
        }

    def test_a_value_may_contain_a_colon(self) -> None:
        assert parse_params(["rss:url=https://x.test/f"])["rss"]["url"].startswith("https:")

    def test_a_malformed_pair_says_what_it_wanted(self) -> None:
        with pytest.raises(SystemExit, match="SOURCE:NAME=VALUE"):
            parse_params(["nonsense"])


class TestFreePort:
    def test_it_finds_something(self) -> None:
        assert 1024 < free_port("127.0.0.1", 8765) < 70000

    def test_it_steps_past_a_port_in_use(self) -> None:
        import socket

        with socket.socket() as taken:
            taken.bind(("127.0.0.1", 0))
            taken.listen(1)
            busy = taken.getsockname()[1]
            assert free_port("127.0.0.1", busy) != busy


# ------------------------------------------------------------------ sources


class TestSourcesCommand:
    def test_it_lists_what_is_installed(
        self, home: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        assert run_cli(home, "sources") == 0
        assert "clifake" in capsys.readouterr().out

    def test_json_output_parses(self, home: Path, capsys: pytest.CaptureFixture[str]) -> None:
        run_cli(home, "sources", "--json")
        listed = json.loads(capsys.readouterr().out)
        assert {source["id"] for source in listed} >= {"rss", "hzz"}


# ---------------------------------------------------------------------- run


class TestRunCommand:
    def test_a_search_saves_what_it_found(self, home: Path) -> None:
        CliSource.postings = [ad(1), ad(2)]
        assert run_cli(home, "run", "clifake") == 0
        assert len(stored(home)) == 2

    def test_no_write_leaves_the_database_alone(self, home: Path) -> None:
        CliSource.postings = [ad(1)]
        assert run_cli(home, "run", "clifake", "--no-write") == 0
        assert stored(home) == []

    def test_an_unknown_source_names_the_alternatives(
        self, home: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        assert run_cli(home, "run", "definitely-not-a-source") == 2
        assert "rss" in capsys.readouterr().err

    def test_parameters_reach_the_source(self, home: Path) -> None:
        CliSource.postings = [ad(1)]
        run_cli(home, "run", "clifake", "--param", "clifake:county=8")
        assert CliSource.seen_params == {"county": "8"}

    def test_a_saved_profile_is_applied(self, home: Path) -> None:
        home.mkdir(parents=True, exist_ok=True)
        with Database(home / "jobsheet.sqlite3") as db:
            db.save_profile(
                "gis", "search", {"keyword_groups": [{"name": "GIS", "terms": ["gis"]}]}
            )

        CliSource.postings = [ad(1), ad(2, title="Pastry chef", description="")]
        run_cli(home, "run", "clifake", "--profile", "gis")

        assert [row.posting.title for row in stored(home)] == ["GIS Engineer 1"]

    def test_a_missing_profile_is_an_error_not_a_silent_default(
        self, home: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        assert run_cli(home, "run", "clifake", "--profile", "never-saved") == 2
        assert "never-saved" in capsys.readouterr().err

    def test_the_run_is_recorded(self, home: Path) -> None:
        CliSource.postings = [ad(1)]
        run_cli(home, "run", "clifake")

        with Database(home / "jobsheet.sqlite3") as db:
            (record,) = db.runs()
            assert record["added"] == 1

    def test_source_health_is_recorded(self, home: Path) -> None:
        CliSource.postings = [ad(1)]
        run_cli(home, "run", "clifake")

        with Database(home / "jobsheet.sqlite3") as db:
            (health,) = db.source_health()
            assert health["source_id"] == "clifake"
            assert health["last_count"] == 1

    def test_a_second_run_does_not_re_add_the_same_ad(self, home: Path) -> None:
        CliSource.postings = [ad(1)]
        run_cli(home, "run", "clifake")
        run_cli(home, "run", "clifake")
        assert len(stored(home)) == 1


# ------------------------------------------------------------------- export


class TestExportWorkbook:
    def test_it_writes_a_workbook(self, home: Path) -> None:
        CliSource.postings = [ad(1), ad(2)]
        run_cli(home, "run", "clifake")

        assert run_cli(home, "export") == 0
        assert (home / "jobs.xlsx").exists()

    def test_the_workbook_can_go_wherever_the_user_wants(
        self, home: Path, tmp_path: Path
    ) -> None:
        CliSource.postings = [ad(1)]
        run_cli(home, "run", "clifake")

        desktop = tmp_path / "Desktop" / "my jobs.xlsx"
        assert main(["--home", str(home), "--workbook", str(desktop), "export"]) == 0
        assert desktop.exists()

    def test_an_open_workbook_fails_rather_than_appearing_to_work(
        self, home: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        CliSource.postings = [ad(1)]
        run_cli(home, "run", "clifake")
        run_cli(home, "export")

        (home / "~$jobs.xlsx").write_bytes(b"excel has this open")

        assert run_cli(home, "export") == 1
        assert "Excel" in capsys.readouterr().err

    def test_edits_made_in_excel_come_back_before_the_rewrite(self, home: Path) -> None:
        """The step that, left out, once erased 88 hand-placed ticks."""
        CliSource.postings = [ad(1)]
        run_cli(home, "run", "clifake")
        run_cli(home, "export")

        path = home / "jobs.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        sheet.cell(row=2, column=headers.index("Status") + 1, value="applied")
        workbook.save(path)
        workbook.close()

        assert run_cli(home, "export") == 0
        assert stored(home)[0].status is ApplicationStatus.APPLIED

    def test_a_saved_layout_is_used(self, home: Path) -> None:
        CliSource.postings = [ad(1)]
        run_cli(home, "run", "clifake")

        mine = SheetLayout(
            columns=[ColumnSpec(key="title", label="Role"), ColumnSpec(key="company", label="Firm")]
        )
        with Database(home / "jobsheet.sqlite3") as db:
            db.save_profile("mine", "layout", mine.model_dump(mode="json"))

        assert run_cli(home, "export", "--layout", "mine") == 0

        workbook = openpyxl.load_workbook(home / "jobs.xlsx")
        assert [cell.value for cell in workbook.active[1]] == ["Role", "Firm"]
        workbook.close()

    def test_a_missing_layout_is_an_error(self, home: Path) -> None:
        assert run_cli(home, "export", "--layout", "never-saved") == 2

    def test_a_status_filter_narrows_what_is_written(self, home: Path) -> None:
        CliSource.postings = [ad(1), ad(2)]
        run_cli(home, "run", "clifake")

        with Database(home / "jobsheet.sqlite3") as db:
            # Through the tracker, not `save_row`: a status is the user's, and
            # saving a row deliberately cannot set one.
            Tracker(db).set_status(db.all_rows()[0].dedup_key, ApplicationStatus.SKIPPED)

        run_cli(home, "export", "--status", "skipped")

        workbook = openpyxl.load_workbook(home / "jobs.xlsx")
        assert workbook.active.max_row == 2  # header plus one job
        workbook.close()


class TestExportData:
    def test_csv_goes_to_standard_output(
        self, home: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        CliSource.postings = [ad(1)]
        run_cli(home, "run", "clifake")

        assert run_cli(home, "export", "csv") == 0
        out = capsys.readouterr().out
        assert "GIS Engineer 1" in out

    def test_csv_can_go_to_a_file(self, home: Path, tmp_path: Path) -> None:
        CliSource.postings = [ad(1)]
        run_cli(home, "run", "clifake")

        target = tmp_path / "out" / "jobs.csv"
        assert run_cli(home, "export", "csv", "--out", str(target)) == 0
        assert "GIS Engineer 1" in target.read_text(encoding="utf-8")

    def test_json_parses(self, home: Path, capsys: pytest.CaptureFixture[str]) -> None:
        CliSource.postings = [ad(1), ad(2)]
        run_cli(home, "run", "clifake")

        run_cli(home, "export", "json")
        records = json.loads(capsys.readouterr().out)
        assert len(records) == 2
        assert records[0]["dedup_key"]

    def test_jsonl_is_one_record_per_line(
        self, home: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        CliSource.postings = [ad(1), ad(2)]
        run_cli(home, "run", "clifake")

        run_cli(home, "export", "jsonl")
        lines = [line for line in capsys.readouterr().out.splitlines() if line]
        assert len(lines) == 2
        assert all(json.loads(line)["title"] for line in lines)

    def test_csv_follows_the_workbooks_own_layout(self, home: Path, tmp_path: Path) -> None:
        """Both exports of the same search should have the same headers."""
        CliSource.postings = [ad(1)]
        run_cli(home, "run", "clifake")
        run_cli(home, "export")

        target = tmp_path / "jobs.csv"
        run_cli(home, "export", "csv", "--out", str(target))

        header = target.read_text(encoding="utf-8").splitlines()[0].lstrip("﻿")
        assert header.split(",")[:2] == [
            column.label for column in default_layout().columns[:2]
        ]
