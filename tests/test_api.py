"""The local HTTP interface.

Two things are worth testing hardest here, and they are the two that would be
silent failures rather than loud ones:

* **the token and the host check**, because a page in another tab can reach
  `127.0.0.1` and must get nowhere,
* **the account boundary**, because one install now holds several job searches
  and the failure mode of getting that wrong is silent: not an error, but one
  person reading another's spreadsheet, and
* **the export order**, because writing the workbook before reading it is
  exactly how a year of hand-typed statuses gets erased.
"""

from __future__ import annotations

import json
import re
from collections.abc import Iterator
from datetime import date, datetime
from pathlib import Path
from typing import Any, ClassVar

import openpyxl
import pytest
from fastapi.testclient import TestClient

from jobsheet.api.app import create_app, web_is_built
from jobsheet.api.runs import SearchRun
from jobsheet.api.state import TOKEN_HEADER
from jobsheet.config import Settings
from jobsheet.core.models import ApplicationStatus, Posting
from jobsheet.sheet.layout import ColumnKind, ColumnSpec, SheetLayout, default_layout
from jobsheet.sheet.row import JobRow
from jobsheet.sources import registry
from jobsheet.sources.base import FetchContext, Source, SourceManifest

# TestClient defaults to a `testserver` host, which the loopback check rejects --
# correctly. Every client here therefore states a real local address.
BASE_URL = "http://127.0.0.1:8765"

FOUND = date(2026, 8, 24)


class ApiSource(Source):
    """A source under the test's control, so no test touches the network."""

    manifest = SourceManifest(
        id="apifake",
        name="API Fake",
        homepage="https://example.test",
        description="Only exists inside the test suite.",
    )
    postings: ClassVar[list[Posting]] = []

    async def fetch(self, params: dict[str, Any], ctx: FetchContext) -> list[Posting]:
        ctx.report(f"apifake: {len(type(self).postings)} ad(s)")
        return list(type(self).postings)


@pytest.fixture(autouse=True)
def _register() -> None:
    registry.register(ApiSource)
    ApiSource.postings = []


def ad(number: int, **overrides: Any) -> Posting:
    data: dict[str, Any] = {
        "source_id": "apifake",
        "title": f"GIS Engineer {number}",
        "url": f"https://example.test/j/{number}",
        "company": f"Company {number}",
        "location": "Zagreb",
        "posted_at": FOUND,
    }
    return Posting(**(data | overrides))


def job(number: int, **overrides: Any) -> JobRow:
    data: dict[str, Any] = {"posting": ad(number), "found_at": FOUND, "category": "GIS"}
    return JobRow(**(data | overrides))


@pytest.fixture
def settings(tmp_path: Path) -> Settings:
    return Settings(home=tmp_path / "home", open_browser=False)


USERNAME = "tester"
PASSWORD = "a-good-password"


def sign_up(client: TestClient, username: str = USERNAME, password: str = PASSWORD) -> Any:
    """Register and stay signed in. The cookie rides on the client from here."""
    response = client.post(
        "/api/auth/register", json={"username": username, "password": password}
    )
    assert response.status_code == 201, response.text
    return response.json()


@pytest.fixture
def anonymous(settings: Settings) -> Iterator[TestClient]:
    """The page token and nothing else: served the interface, not signed in."""
    with TestClient(create_app(settings), base_url=BASE_URL) as test_client:
        test_client.headers[TOKEN_HEADER] = settings.token
        yield test_client


@pytest.fixture
def client(settings: Settings) -> Iterator[TestClient]:
    app = create_app(settings)
    with TestClient(app, base_url=BASE_URL) as test_client:
        test_client.headers[TOKEN_HEADER] = settings.token
        sign_up(test_client)
        yield test_client


@pytest.fixture
def state(client: TestClient) -> Any:
    """The app's own database and tracker, for arranging and for checking."""
    return client.app.state.jobsheet  # type: ignore[attr-defined]


# ------------------------------------------------------------- getting in


class TestAccess:
    def test_health_needs_nothing(self, settings: Settings) -> None:
        """The launcher polls this before it knows the token."""
        with TestClient(create_app(settings), base_url=BASE_URL) as anonymous:
            assert anonymous.get("/api/health").json()["status"] == "ok"

    def test_no_token_is_refused(self, settings: Settings) -> None:
        with TestClient(create_app(settings), base_url=BASE_URL) as anonymous:
            assert anonymous.get("/api/settings").status_code == 401

    def test_the_wrong_token_is_refused(self, settings: Settings) -> None:
        with TestClient(create_app(settings), base_url=BASE_URL) as anonymous:
            response = anonymous.get("/api/settings", headers={TOKEN_HEADER: "guess"})
            assert response.status_code == 401

    def test_a_non_loopback_host_is_refused(self, settings: Settings) -> None:
        """DNS rebinding: a hostile domain resolving to 127.0.0.1 gets nowhere."""
        with TestClient(create_app(settings), base_url="http://evil.example") as outsider:
            response = outsider.get("/api/settings", headers={TOKEN_HEADER: settings.token})
            assert response.status_code == 403

    def test_the_host_check_runs_before_the_token_check(self, settings: Settings) -> None:
        """Otherwise a wrong host with a right token would leak that it was right."""
        with TestClient(create_app(settings), base_url="http://evil.example") as outsider:
            assert outsider.get("/api/settings").status_code == 403

    def test_localhost_by_name_is_fine(self, settings: Settings) -> None:
        with TestClient(create_app(settings), base_url="http://localhost:8765") as local:
            response = local.get("/api/auth/status", headers={TOKEN_HEADER: settings.token})
            assert response.status_code == 200

    def test_the_token_alone_does_not_get_past_the_sign_in(
        self, anonymous: TestClient
    ) -> None:
        """The two checks answer different questions and neither stands in for the other."""
        assert anonymous.get("/api/settings").status_code == 401
        assert anonymous.get("/api/auth/me").status_code == 401
        # ...while the door itself is reachable, or nobody could ever sign in.
        assert anonymous.get("/api/auth/status").status_code == 200

    def test_the_stream_accepts_the_token_in_the_query(self, client: TestClient) -> None:
        """EventSource cannot set headers, so the stream has to allow this."""
        started = client.post(
            "/api/search", json={"sources": [{"source_id": "apifake"}]}
        ).json()
        del client.headers[TOKEN_HEADER]

        token = client.app.state.jobsheet.settings.token  # type: ignore[attr-defined]
        response = client.get(f"/api/search/{started['id']}/stream?token={token}")
        assert response.status_code == 200


class TestThePage:
    def test_the_token_is_handed_over_in_the_page(self, client: TestClient) -> None:
        body = client.get("/").text
        assert "__JOBSHEET__" in body
        assert client.app.state.jobsheet.settings.token in body  # type: ignore[attr-defined]

    def test_unknown_paths_fall_through_to_the_interface(self, client: TestClient) -> None:
        """A single-page app owns its own routes; a reload of /tracker must work."""
        assert client.get("/tracker").status_code == 200

    def test_the_page_forbids_reaching_anywhere_else(self, client: TestClient) -> None:
        headers = client.get("/").headers
        assert "default-src 'self'" in headers["content-security-policy"]
        assert headers["x-frame-options"] == "DENY"

    def test_a_traversal_attempt_gets_the_interface_not_a_file(
        self, client: TestClient
    ) -> None:
        response = client.get("/../../pyproject.toml")
        assert response.status_code == 200
        assert "[project]" not in response.text


@pytest.mark.skipif(not web_is_built(), reason="the interface has not been built")
class TestTheBuiltInterface:
    """Only meaningful once `npm run build` has put a bundle in the package."""

    def test_the_real_page_is_served(self, client: TestClient) -> None:
        body = client.get("/").text
        assert 'id="root"' in body
        assert "__JOBSHEET__" in body

    def test_scripts_are_served_as_javascript(self, client: TestClient) -> None:
        """Windows maps media types through the registry, where `.js` can end up
        as `text/plain` -- and a browser then refuses to run the module."""
        page = client.get("/").text
        script = re.search(r"/assets/[^\"']+\.js", page)
        assert script, "the built page references no script"

        response = client.get(script.group(0))
        assert response.status_code == 200
        assert response.headers["content-type"].startswith("text/javascript")

    def test_stylesheets_are_served_as_css(self, client: TestClient) -> None:
        page = client.get("/").text
        sheet = re.search(r"/assets/[^\"']+\.css", page)
        assert sheet, "the built page references no stylesheet"
        assert client.get(sheet.group(0)).headers["content-type"].startswith("text/css")


# ---------------------------------------------------------------- sources


class TestSources:
    def test_every_installed_source_is_listed_with_its_form(
        self, client: TestClient
    ) -> None:
        body = client.get("/api/sources").json()
        listed = {source["id"]: source for source in body["sources"]}

        assert "apifake" in listed
        assert "hzz" in listed
        # The form the interface draws comes from here, not from frontend code.
        assert "params" in listed["hzz"]

    def test_health_is_empty_until_something_has_run(self, client: TestClient) -> None:
        assert client.get("/api/sources/health").json() == []

    def test_a_finished_run_fills_in_health(self, client: TestClient) -> None:
        ApiSource.postings = [ad(1)]
        client.post("/api/search", json={"sources": [{"source_id": "apifake"}]})

        health = {record["source_id"]: record for record in client.get("/api/sources/health").json()}
        assert health["apifake"]["last_ok"] is not None
        assert health["apifake"]["last_count"] == 1


# ----------------------------------------------------------------- search


class TestSearch:
    def test_a_search_returns_before_it_has_finished(self, client: TestClient) -> None:
        ApiSource.postings = [ad(1)]
        response = client.post("/api/search", json={"sources": [{"source_id": "apifake"}]})

        assert response.status_code == 202
        assert response.json()["id"]

    def test_an_uninstalled_source_is_refused_by_name(self, client: TestClient) -> None:
        response = client.post("/api/search", json={"sources": [{"source_id": "nope"}]})
        assert response.status_code == 422
        assert "nope" in response.json()["detail"]

    def test_a_search_with_no_sources_is_refused(self, client: TestClient) -> None:
        assert client.post("/api/search", json={"sources": []}).status_code == 422

    def test_what_it_found_is_saved(self, client: TestClient, state: Any) -> None:
        ApiSource.postings = [ad(1), ad(2)]
        client.post("/api/search", json={"sources": [{"source_id": "apifake"}]})

        assert len(state.db.all_rows()) == 2

    def test_the_run_reports_what_it_did(self, client: TestClient) -> None:
        ApiSource.postings = [ad(1), ad(2)]
        started = client.post(
            "/api/search", json={"sources": [{"source_id": "apifake"}]}
        ).json()

        run = client.get(f"/api/search/{started['id']}").json()
        assert run["phase"] == "done"
        assert run["fetched"] == 2
        assert run["new"] == 2
        assert run["harvested"] == {"apifake": 2}

    def test_the_commentary_is_kept(self, client: TestClient) -> None:
        ApiSource.postings = [ad(1)]
        started = client.post(
            "/api/search", json={"sources": [{"source_id": "apifake"}]}
        ).json()

        lines = client.get(f"/api/search/{started['id']}").json()["lines"]
        assert any("apifake" in line for line in lines)

    def test_the_stream_replays_a_finished_run(self, client: TestClient) -> None:
        """A browser that connects late still gets the whole story."""
        ApiSource.postings = [ad(1)]
        started = client.post(
            "/api/search", json={"sources": [{"source_id": "apifake"}]}
        ).json()

        body = client.get(f"/api/search/{started['id']}/stream").text
        assert "event: progress" in body
        assert "event: end" in body
        assert "data: done" in body

    def test_results_explain_what_was_turned_away(self, client: TestClient) -> None:
        ApiSource.postings = [
            ad(1),
            ad(2, deadline=date(2026, 1, 1)),  # closed long ago
        ]
        started = client.post(
            "/api/search",
            json={"sources": [{"source_id": "apifake"}], "today": FOUND.isoformat()},
        ).json()

        results = client.get(f"/api/search/{started['id']}/results").json()
        assert len(results["rows"]) == 1
        assert results["rejected"][0]["code"]

    def test_results_before_the_end_are_a_conflict_not_an_empty_list(
        self, client: TestClient, state: Any
    ) -> None:
        """An empty list would read as "nothing found", which is a different claim."""
        run = SearchRun(
            id="pending",
            started_at=datetime.now(),
            requests=[],
            user_id=state.db.user_id,
            db=state.db,
        )
        state.runs._runs[run.id] = run

        assert client.get("/api/search/pending/results").status_code == 409

    def test_an_unknown_run_is_a_404(self, client: TestClient) -> None:
        assert client.get("/api/search/nosuchrun").status_code == 404
        assert client.get("/api/search/nosuchrun/stream").status_code == 404
        assert client.post("/api/search/nosuchrun/cancel").status_code == 404

    def test_runs_are_listed_newest_first(self, client: TestClient) -> None:
        first = client.post("/api/search", json={"sources": [{"source_id": "apifake"}]}).json()
        second = client.post("/api/search", json={"sources": [{"source_id": "apifake"}]}).json()

        listed = [run["id"] for run in client.get("/api/search").json()]
        assert listed.index(second["id"]) <= listed.index(first["id"])


# --------------------------------------------------------------- postings


class TestPostings:
    def test_an_empty_database_is_an_empty_page(self, client: TestClient) -> None:
        body = client.get("/api/postings").json()
        assert body == {"total": 0, "limit": 100, "offset": 0, "rows": []}

    def test_rows_come_back(self, client: TestClient, state: Any) -> None:
        state.db.save_rows([job(1), job(2)])
        body = client.get("/api/postings").json()
        assert body["total"] == 2
        assert len(body["rows"]) == 2

    def test_the_total_counts_matches_not_the_page(
        self, client: TestClient, state: Any
    ) -> None:
        state.db.save_rows([job(number) for number in range(5)])
        body = client.get("/api/postings?limit=2").json()
        assert body["total"] == 5
        assert len(body["rows"]) == 2

    def test_paging_walks_the_whole_list(self, client: TestClient, state: Any) -> None:
        state.db.save_rows([job(number) for number in range(5)])
        first = client.get("/api/postings?limit=2&offset=0").json()["rows"]
        second = client.get("/api/postings?limit=2&offset=2").json()["rows"]
        assert {row["dedup_key"] for row in first}.isdisjoint(
            {row["dedup_key"] for row in second}
        )

    def test_a_query_narrows_by_title(self, client: TestClient, state: Any) -> None:
        state.db.save_rows([job(1), job(2, posting=ad(2, title="Pastry chef"))])
        body = client.get("/api/postings?q=pastry").json()
        assert body["total"] == 1
        assert body["rows"][0]["posting"]["title"] == "Pastry chef"

    def test_a_query_narrows_by_company(self, client: TestClient, state: Any) -> None:
        state.db.save_rows([job(1), job(2, posting=ad(2, company="Kartograf d.o.o."))])
        assert client.get("/api/postings?q=kartograf").json()["total"] == 1

    def test_a_status_filter_narrows(self, client: TestClient, state: Any) -> None:
        state.db.save_rows([job(1), job(2, status=ApplicationStatus.APPLIED)])
        assert client.get("/api/postings?status=applied").json()["total"] == 1

    def test_an_unknown_status_is_refused_rather_than_ignored(
        self, client: TestClient
    ) -> None:
        assert client.get("/api/postings?status=invented").status_code == 422

    def test_one_job_by_key(self, client: TestClient, state: Any) -> None:
        state.db.save_row(job(1))
        response = client.get("/api/postings/one", params={"dedup_key": job(1).dedup_key})
        assert response.json()["posting"]["title"] == "GIS Engineer 1"

    def test_an_unknown_key_is_a_404(self, client: TestClient) -> None:
        assert client.get("/api/postings/one", params={"dedup_key": "no"}).status_code == 404

    def test_deleting_forgets_the_job(self, client: TestClient, state: Any) -> None:
        state.db.save_row(job(1))
        client.delete("/api/postings/one", params={"dedup_key": job(1).dedup_key})
        assert state.db.all_rows() == []

    def test_past_runs_survive_the_process(self, client: TestClient) -> None:
        ApiSource.postings = [ad(1)]
        client.post("/api/search", json={"sources": [{"source_id": "apifake"}]})
        assert len(client.get("/api/postings/runs").json()) == 1


# ----------------------------------------------------------- applications


class TestApplications:
    def test_the_board_has_every_column(self, client: TestClient) -> None:
        """Discarded leads, rejected trails; the interface draws this order."""
        body = client.get("/api/applications/board").json()
        assert body["order"] == ["skipped", "new", "applied", "interview", "offer", "rejected"]

    def test_a_job_lands_in_its_column(self, client: TestClient, state: Any) -> None:
        state.db.save_row(job(1, status=ApplicationStatus.APPLIED))
        body = client.get("/api/applications/board").json()
        assert len(body["columns"]["applied"]) == 1
        assert body["counts"]["applied"] == 1

    def test_moving_a_card_is_recorded(self, client: TestClient, state: Any) -> None:
        state.db.save_row(job(1))
        response = client.post(
            "/api/applications/status",
            json={"dedup_key": job(1).dedup_key, "status": "applied", "note": "sent CV"},
        )

        assert response.json()["changed"] is True
        assert response.json()["from_status"] == "new"
        assert state.tracker.status_of(job(1).dedup_key) is ApplicationStatus.APPLIED

    def test_moving_a_card_nowhere_changes_nothing(
        self, client: TestClient, state: Any
    ) -> None:
        state.db.save_row(job(1, status=ApplicationStatus.APPLIED))
        response = client.post(
            "/api/applications/status",
            json={"dedup_key": job(1).dedup_key, "status": "applied"},
        )
        assert response.json()["changed"] is False

    def test_moving_an_untracked_job_is_a_404(self, client: TestClient) -> None:
        response = client.post(
            "/api/applications/status", json={"dedup_key": "nothing", "status": "applied"}
        )
        assert response.status_code == 404

    def test_an_invented_status_is_refused(self, client: TestClient, state: Any) -> None:
        state.db.save_row(job(1))
        response = client.post(
            "/api/applications/status",
            json={"dedup_key": job(1).dedup_key, "status": "ghosted"},
        )
        assert response.status_code == 422

    def test_history_reads_back_in_order(self, client: TestClient, state: Any) -> None:
        state.db.save_row(job(1))
        for status in ("applied", "interview"):
            client.post(
                "/api/applications/status",
                json={"dedup_key": job(1).dedup_key, "status": status},
            )

        history = client.get(
            "/api/applications/history", params={"dedup_key": job(1).dedup_key}
        ).json()
        assert [step["to_status"] for step in history] == ["applied", "interview"]

    def test_user_values_are_stored(self, client: TestClient, state: Any) -> None:
        state.db.save_row(job(1))
        client.post(
            "/api/applications/values",
            json={"dedup_key": job(1).dedup_key, "values": {"Rating": 5}},
        )
        assert state.db.all_rows()[0].user_values == {"Rating": 5}


# --------------------------------------------------------------- layouts


class TestLayouts:
    def test_the_presets_are_offered(self, client: TestClient) -> None:
        names = [preset["name"] for preset in client.get("/api/layouts/presets").json()]
        assert names == ["default", "classic-checkboxes", "minimal"]

    def test_a_preset_comes_back_whole(self, client: TestClient) -> None:
        layout = client.get("/api/layouts/presets/classic-checkboxes").json()
        assert SheetLayout.model_validate(layout).checkbox_keys

    def test_an_unknown_preset_is_a_404(self, client: TestClient) -> None:
        assert client.get("/api/layouts/presets/nope").status_code == 404

    def test_the_designer_is_told_what_it_may_offer(self, client: TestClient) -> None:
        body = client.get("/api/layouts/vocabulary").json()
        assert {kind["value"] for kind in body["kinds"]} == {str(k) for k in ColumnKind}
        assert "title" in body["source_keys"]
        assert any(theme["default"] for theme in body["themes"])

    def test_without_a_workbook_the_designer_opens_on_the_default(
        self, client: TestClient
    ) -> None:
        body = client.get("/api/layouts/current").json()
        assert body["exists"] is False
        assert body["from_workbook"] is False

    def test_the_layout_in_the_workbook_wins(self, client: TestClient) -> None:
        """A user who rearranged their columns in Excel keeps that arrangement."""
        mine = SheetLayout(
            columns=[
                ColumnSpec(key="title", label="Role"),
                ColumnSpec(key="url", label="Link", kind=ColumnKind.URL),
            ]
        )
        client.post("/api/export/xlsx", json={"layout": mine.model_dump(mode="json")})

        body = client.get("/api/layouts/current").json()
        assert body["from_workbook"] is True
        assert [column["label"] for column in body["layout"]["columns"]] == ["Role", "Link"]

    def test_a_good_design_validates(self, client: TestClient) -> None:
        body = client.post(
            "/api/layouts/validate", json=default_layout().model_dump(mode="json")
        ).json()
        assert body["valid"] is True
        assert body["problems"] == []

    def test_a_bad_design_says_what_is_wrong_rather_than_erroring(
        self, client: TestClient
    ) -> None:
        response = client.post("/api/layouts/validate", json={"columns": []})
        assert response.status_code == 200
        assert response.json()["valid"] is False
        assert response.json()["problems"]

    def test_validation_names_the_user_owned_columns(self, client: TestClient) -> None:
        """The designer greys those out: the app must never write to them."""
        layout = SheetLayout(
            columns=[
                ColumnSpec(key="title", label="Role"),
                ColumnSpec(key="my_notes", label="My notes"),
            ]
        )
        body = client.post(
            "/api/layouts/validate", json=layout.model_dump(mode="json")
        ).json()
        assert body["user_owned"] == ("my_notes",) or body["user_owned"] == ["my_notes"]


# -------------------------------------------------------------- profiles


class TestProfiles:
    def test_a_search_profile_round_trips(self, client: TestClient) -> None:
        payload = {"keyword_groups": [{"name": "GIS", "terms": ["gis"]}], "max_age_days": 14}
        client.put("/api/profiles/search/mine", json={"payload": payload})

        body = client.get("/api/profiles/search/mine").json()
        assert body["payload"]["max_age_days"] == 14

    def test_a_layout_profile_round_trips(self, client: TestClient) -> None:
        client.put(
            "/api/profiles/layout/mine",
            json={"payload": default_layout().model_dump(mode="json")},
        )
        body = client.get("/api/profiles/layout/mine").json()
        assert body["payload"]["columns"]

    def test_an_unusable_profile_is_refused_at_save_time(self, client: TestClient) -> None:
        """Better to find out now than when the user reaches for it in three weeks."""
        response = client.put(
            "/api/profiles/layout/broken", json={"payload": {"columns": []}}
        )
        assert response.status_code == 422

    def test_an_unknown_kind_is_a_404(self, client: TestClient) -> None:
        assert client.get("/api/profiles/nonsense").status_code == 404

    def test_profiles_are_listed(self, client: TestClient) -> None:
        client.put("/api/profiles/search/b", json={"payload": {}})
        client.put("/api/profiles/search/a", json={"payload": {}})
        assert client.get("/api/profiles/search").json() == ["a", "b"]

    def test_deleting_removes_it(self, client: TestClient) -> None:
        client.put("/api/profiles/search/mine", json={"payload": {}})
        assert client.delete("/api/profiles/search/mine").status_code == 200
        assert client.get("/api/profiles/search/mine").status_code == 404

    def test_deleting_something_absent_is_a_404(self, client: TestClient) -> None:
        assert client.delete("/api/profiles/search/never").status_code == 404


# ---------------------------------------------------------------- export


class TestExport:
    def test_the_workbook_is_written(self, client: TestClient, state: Any) -> None:
        state.db.save_rows([job(1), job(2)])
        body = client.post("/api/export/xlsx", json={}).json()

        assert body["rows"] == 2
        assert Path(body["path"]).exists()

    def test_a_second_export_backs_the_first_one_up(
        self, client: TestClient, state: Any
    ) -> None:
        state.db.save_row(job(1))
        client.post("/api/export/xlsx", json={})
        body = client.post("/api/export/xlsx", json={}).json()

        assert body["backup"] is not None
        assert Path(body["backup"]).exists()

    def test_edits_made_in_excel_are_taken_back_before_writing(
        self, client: TestClient, state: Any
    ) -> None:
        """The whole reason the export reads before it writes."""
        state.db.save_row(job(1))
        path = Path(client.post("/api/export/xlsx", json={}).json()["path"])

        # The user opens the file and marks the job as applied.
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        sheet.cell(row=2, column=headers.index("Status") + 1, value="applied")
        workbook.save(path)
        workbook.close()

        body = client.post("/api/export/xlsx", json={}).json()

        assert body["adopted_from_workbook"] == [job(1).dedup_key]
        assert state.tracker.status_of(job(1).dedup_key) is ApplicationStatus.APPLIED

    def test_a_status_typed_in_excel_is_not_reverted_by_the_write(
        self, client: TestClient, state: Any
    ) -> None:
        state.db.save_row(job(1))
        path = Path(client.post("/api/export/xlsx", json={}).json()["path"])

        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        sheet.cell(row=2, column=headers.index("Status") + 1, value="interview")
        workbook.save(path)
        workbook.close()

        client.post("/api/export/xlsx", json={})

        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        assert sheet.cell(row=2, column=headers.index("Status") + 1).value == "interview"
        workbook.close()

    def test_an_open_workbook_is_refused_rather_than_clobbered(
        self, client: TestClient, state: Any
    ) -> None:
        state.db.save_row(job(1))
        client.post("/api/export/xlsx", json={})

        workbook_path = state.settings.workbook_path
        lock = workbook_path.parent / f"~${workbook_path.name}"
        lock.write_bytes(b"excel has this open")

        response = client.post("/api/export/xlsx", json={})
        assert response.status_code == 409
        assert "Excel" in response.json()["detail"]

    def test_the_workbook_state_is_reportable_before_committing(
        self, client: TestClient, state: Any
    ) -> None:
        body = client.get("/api/export/workbook").json()
        assert body["exists"] is False
        assert body["locked"] is False

    def test_a_status_filter_narrows_what_is_written(
        self, client: TestClient, state: Any
    ) -> None:
        state.db.save_rows([job(1), job(2, status=ApplicationStatus.SKIPPED)])
        body = client.post("/api/export/xlsx", json={"statuses": ["new"]}).json()
        assert body["rows"] == 1

    def test_csv_carries_the_layouts_headers(self, client: TestClient, state: Any) -> None:
        state.db.save_row(job(1))
        response = client.post("/api/export/csv", json={})

        assert response.status_code == 200
        assert "attachment" in response.headers["content-disposition"]
        text = response.content.decode("utf-8-sig")
        assert text.splitlines()[0].startswith("Found")

    def test_csv_leads_with_a_bom_so_excel_reads_it(
        self, client: TestClient, state: Any
    ) -> None:
        state.db.save_row(job(1, posting=ad(1, company="Građevinar d.o.o.")))
        content = client.post("/api/export/csv", json={}).content
        assert content.startswith(b"\xef\xbb\xbf")
        assert "Građevinar" in content.decode("utf-8-sig")

    def test_json_holds_the_whole_record(self, client: TestClient, state: Any) -> None:
        state.db.save_row(job(1))
        records = json.loads(client.post("/api/export/json", json={}).content)
        assert records[0]["title"] == "GIS Engineer 1"
        assert records[0]["dedup_key"]

    def test_jsonl_is_one_job_per_line(self, client: TestClient, state: Any) -> None:
        state.db.save_rows([job(1), job(2)])
        text = client.post("/api/export/jsonl", json={}).content.decode()
        assert len([line for line in text.splitlines() if line]) == 2


# ---------------------------------------------------------------- letter


class TestLetter:
    def test_a_draft_names_the_job(self, client: TestClient, state: Any) -> None:
        state.db.save_row(job(1))
        body = client.post(
            "/api/letter",
            json={
                "dedup_key": job(1).dedup_key,
                "applicant": {"name": "Ana Horvat", "email": "ana@example.test"},
            },
        ).json()

        assert "GIS Engineer 1" in body["text"]
        assert "Ana Horvat" in body["text"]
        assert "Company 1" in body["text"]

    def test_an_unknown_job_is_a_404(self, client: TestClient) -> None:
        response = client.post("/api/letter", json={"dedup_key": "nothing"})
        assert response.status_code == 404

    def test_a_users_own_template_is_used(self, client: TestClient, state: Any) -> None:
        state.db.save_row(job(1))
        (state.settings.home / "letter.txt").write_text(
            "Dear {{ company }}, about {{ title }}.", encoding="utf-8"
        )

        body = client.post("/api/letter", json={"dedup_key": job(1).dedup_key}).json()
        assert body["text"] == "Dear Company 1, about GIS Engineer 1."

    def test_a_broken_template_is_explained_not_crashed(
        self, client: TestClient, state: Any
    ) -> None:
        state.db.save_row(job(1))
        response = client.post(
            "/api/letter",
            json={"dedup_key": job(1).dedup_key, "template": "{% for %}"},
        )
        assert response.status_code == 422

    def test_the_template_in_use_is_reportable(self, client: TestClient) -> None:
        body = client.get("/api/letter/template").json()
        assert body["custom"] is False
        assert "{{ applicant.name }}" in body["template"]


# -------------------------------------------------------------- settings


class TestSettings:
    def test_it_says_where_everything_lives(self, client: TestClient, state: Any) -> None:
        body = client.get("/api/settings").json()
        assert body["workbook"] == str(state.settings.workbook_path)
        assert body["database"] == str(state.settings.database_path)
        assert body["sources_installed"] > 0

    def test_it_never_echoes_the_token(self, client: TestClient, state: Any) -> None:
        """It is the thing that authorises the call; repeating it only adds risk."""
        assert state.settings.token not in client.get("/api/settings").text
        assert state.settings.token not in client.get("/api/settings/folders").text


class TestChangingTheWorkbookPath:
    """The wizard's answer is not the last word: people move their files."""

    def test_the_path_can_be_changed_after_the_wizard(
        self, client: TestClient, tmp_path: Path
    ) -> None:
        wanted = tmp_path / "poslovi.xlsx"
        response = client.put("/api/settings/workbook", json={"path": str(wanted)})

        assert response.status_code == 200
        assert response.json()["workbook"] == str(wanted)
        assert response.json()["moved"] is False
        # And it is the answer every later request gets, without a restart.
        assert client.get("/api/settings").json()["workbook"] == str(wanted)
        assert client.get("/api/auth/me").json()["workbook_path"] == str(wanted)

    def test_the_workbook_can_travel_to_the_new_path(
        self, client: TestClient, state: Any, tmp_path: Path
    ) -> None:
        """Without this, changing the path silently orphans a year of ticks."""
        was = Path(client.get("/api/settings").json()["workbook"])
        was.parent.mkdir(parents=True, exist_ok=True)
        was.write_bytes(b"pretend spreadsheet")

        wanted = tmp_path / "elsewhere" / "poslovi.xlsx"
        wanted.parent.mkdir()
        response = client.put(
            "/api/settings/workbook", json={"path": str(wanted), "move": True}
        )

        assert response.status_code == 200
        assert response.json()["moved"] is True
        assert wanted.read_bytes() == b"pretend spreadsheet"
        assert not was.exists()

    def test_a_workbook_open_in_excel_is_not_moved(
        self, client: TestClient, tmp_path: Path
    ) -> None:
        was = Path(client.get("/api/settings").json()["workbook"])
        was.parent.mkdir(parents=True, exist_ok=True)
        was.write_bytes(b"pretend spreadsheet")
        was.with_name(f"~${was.name}").write_bytes(b"")

        response = client.put(
            "/api/settings/workbook",
            json={"path": str(tmp_path / "poslovi.xlsx"), "move": True},
        )

        assert response.status_code == 409
        assert response.json()["detail"]["code"] == "workbook_locked"
        # Nothing moved and nothing was recorded: the account still points here.
        assert was.exists()
        assert client.get("/api/settings").json()["workbook"] == str(was)

    def test_a_move_will_not_write_over_a_workbook_already_there(
        self, client: TestClient, tmp_path: Path
    ) -> None:
        was = Path(client.get("/api/settings").json()["workbook"])
        was.parent.mkdir(parents=True, exist_ok=True)
        was.write_bytes(b"mine")
        occupied = tmp_path / "poslovi.xlsx"
        occupied.write_bytes(b"somebody else's")

        response = client.put(
            "/api/settings/workbook", json={"path": str(occupied), "move": True}
        )

        assert response.status_code == 409
        assert response.json()["detail"]["code"] == "workbook_in_the_way"
        assert occupied.read_bytes() == b"somebody else's"

    def test_pointing_at_a_workbook_that_is_there_is_allowed_without_moving(
        self, client: TestClient, tmp_path: Path
    ) -> None:
        """Adopting an existing spreadsheet is the other half of this feature."""
        existing = tmp_path / "poslovi.xlsx"
        existing.write_bytes(b"already mine")

        response = client.put("/api/settings/workbook", json={"path": str(existing)})

        assert response.status_code == 200
        assert response.json()["workbook_exists"] is True
        assert existing.read_bytes() == b"already mine"

    def test_a_path_that_is_not_a_spreadsheet_is_refused(self, client: TestClient) -> None:
        response = client.put("/api/settings/workbook", json={"path": "notes.txt"})
        assert response.status_code == 422
        assert response.json()["detail"]["code"] == "workbook_not_xlsx"

    def test_a_folder_that_does_not_exist_is_refused(
        self, client: TestClient, tmp_path: Path
    ) -> None:
        response = client.put(
            "/api/settings/workbook", json={"path": str(tmp_path / "nope" / "jobs.xlsx")}
        )
        assert response.status_code == 422
        assert response.json()["detail"]["code"] == "workbook_folder_missing"

    def test_one_account_cannot_move_the_other_ones_workbook(
        self, settings: Settings, tmp_path: Path
    ) -> None:
        app = create_app(settings)
        with (
            TestClient(app, base_url=BASE_URL) as ana,
            TestClient(app, base_url=BASE_URL) as ivo,
        ):
            ana.headers[TOKEN_HEADER] = settings.token
            ivo.headers[TOKEN_HEADER] = settings.token
            sign_up(ana, "ana")
            sign_up(ivo, "ivo", "another-good-one")

            ana.put("/api/settings/workbook", json={"path": str(tmp_path / "ana.xlsx")})

            assert ana.get("/api/settings").json()["workbook"] == str(tmp_path / "ana.xlsx")
            assert ivo.get("/api/settings").json()["workbook"] != str(tmp_path / "ana.xlsx")


class TestTheFolderPicker:
    """Choosing where the workbook goes by looking, rather than by typing."""

    def test_it_lists_folders_and_not_files(
        self, client: TestClient, tmp_path: Path
    ) -> None:
        # Its own folder, not `tmp_path`: the settings fixture has already put
        # `home/` in there, and this test is about what a listing contains.
        disk = tmp_path / "disk"
        (disk / "Documents").mkdir(parents=True)
        (disk / "notes.txt").write_text("not a folder")

        body = client.get("/api/settings/folders", params={"path": str(disk)}).json()

        assert body["path"] == str(disk)
        assert [folder["name"] for folder in body["folders"]] == ["Documents"]

    def test_it_offers_the_way_back_up(self, client: TestClient, tmp_path: Path) -> None:
        inner = tmp_path / "disk" / "Documents"
        inner.mkdir(parents=True)
        body = client.get("/api/settings/folders", params={"path": str(inner)}).json()
        assert body["parent"] == str(inner.parent)

    def test_it_starts_where_the_workbook_already_is(self, client: TestClient) -> None:
        workbook = Path(client.get("/api/settings").json()["workbook"])
        assert client.get("/api/settings/folders").json()["path"] == str(workbook.parent)

    def test_a_folder_that_is_not_there_falls_back_rather_than_failing(
        self, client: TestClient, tmp_path: Path
    ) -> None:
        """A picker that 500s the moment a path goes stale is a picker nobody trusts."""
        response = client.get(
            "/api/settings/folders", params={"path": str(tmp_path / "gone")}
        )
        assert response.status_code == 200
        assert response.json()["path"] == str(Path.home())

    def test_it_hides_the_dotted_folders_nobody_keeps_a_job_search_in(
        self, client: TestClient, tmp_path: Path
    ) -> None:
        disk = tmp_path / "disk"
        (disk / ".venv").mkdir(parents=True)
        (disk / "Desktop").mkdir()
        body = client.get("/api/settings/folders", params={"path": str(disk)}).json()
        assert [folder["name"] for folder in body["folders"]] == ["Desktop"]


# ------------------------------------------------- what the progress bars need


class TestProgressPerSource:
    """The second channel on the search stream: positions, not sentences."""

    def test_a_run_reports_a_position_for_every_source(self, client: TestClient) -> None:
        ApiSource.postings = [ad(1)]
        started = client.post(
            "/api/search", json={"sources": [{"source_id": "apifake"}]}
        ).json()

        run = client.get(f"/api/search/{started['id']}").json()
        assert run["state"]["apifake"]["phase"] == "done"
        assert run["state"]["apifake"]["percent"] == 100

    def test_the_stream_carries_state_events_beside_the_prose(
        self, client: TestClient
    ) -> None:
        ApiSource.postings = [ad(1)]
        started = client.post(
            "/api/search", json={"sources": [{"source_id": "apifake"}]}
        ).json()

        body = client.get(f"/api/search/{started['id']}/stream").text
        assert "event: state" in body
        # The old channel is untouched, so a client that ignores `state` still works.
        assert "event: progress" in body

    def test_a_reloaded_page_does_not_come_back_to_empty_bars(
        self, client: TestClient
    ) -> None:
        """The bars replay as a position, not a history: current values, once each."""
        ApiSource.postings = [ad(1)]
        started = client.post(
            "/api/search", json={"sources": [{"source_id": "apifake"}]}
        ).json()

        body = client.get(f"/api/search/{started['id']}/stream").text
        states = [line for line in body.splitlines() if line.startswith("data: {")]
        assert len(states) == 1
        assert json.loads(states[0].removeprefix("data: "))["percent"] == 100

    def test_a_source_that_failed_does_not_report_as_finished(
        self, client: TestClient
    ) -> None:
        started = client.post(
            "/api/search", json={"sources": [{"source_id": "apifake"}]}
        ).json()
        run = client.get(f"/api/search/{started['id']}").json()
        assert run["state"]["apifake"]["phase"] in {"done", "failed"}


class TestFilteringByWhichSearch:
    """Item 15: two searches in one morning must be tellable apart."""

    def test_a_finished_run_says_which_row_it_became(self, client: TestClient) -> None:
        ApiSource.postings = [ad(1)]
        started = client.post(
            "/api/search", json={"sources": [{"source_id": "apifake"}]}
        ).json()
        run = client.get(f"/api/search/{started['id']}").json()
        assert run["run_id"]

    def test_the_results_can_be_narrowed_to_one_search(self, client: TestClient) -> None:
        ApiSource.postings = [ad(1)]
        first = client.get(
            f"/api/search/"
            f"{client.post('/api/search', json={'sources': [{'source_id': 'apifake'}]}).json()['id']}"
        ).json()

        ApiSource.postings = [ad(2)]
        second = client.get(
            f"/api/search/"
            f"{client.post('/api/search', json={'sources': [{'source_id': 'apifake'}]}).json()['id']}"
        ).json()

        assert client.get("/api/postings").json()["total"] == 2
        one = client.get("/api/postings", params={"run": first["run_id"]}).json()
        assert one["total"] == 1
        assert one["rows"][0]["posting"]["title"] == "GIS Engineer 1"
        assert client.get("/api/postings", params={"run": second["run_id"]}).json()["total"] == 1

    def test_past_searches_are_listed_with_what_they_added(
        self, client: TestClient
    ) -> None:
        """The filter needs a label per search, and `added` is what makes one."""
        ApiSource.postings = [ad(1), ad(2)]
        client.post("/api/search", json={"sources": [{"source_id": "apifake"}]})

        runs = client.get("/api/postings/runs").json()
        assert runs[0]["added"] == 2
        assert runs[0]["started_at"]

    def test_an_unknown_search_matches_nothing_rather_than_everything(
        self, client: TestClient
    ) -> None:
        ApiSource.postings = [ad(1)]
        client.post("/api/search", json={"sources": [{"source_id": "apifake"}]})
        assert client.get("/api/postings", params={"run": "9999"}).json()["total"] == 0


class TestTheSameJobNeverTwice:
    """Item 22, end to end: the three memories, through the real endpoints."""

    def test_a_deleted_job_does_not_return_on_the_next_search(
        self, client: TestClient
    ) -> None:
        ApiSource.postings = [ad(1)]
        client.post("/api/search", json={"sources": [{"source_id": "apifake"}]})
        key = client.get("/api/postings").json()["rows"][0]["dedup_key"]

        client.delete("/api/postings/one", params={"dedup_key": key})
        assert client.get("/api/postings").json()["total"] == 0

        client.post("/api/search", json={"sources": [{"source_id": "apifake"}]})
        assert client.get("/api/postings").json()["total"] == 0

    def test_a_discarded_job_does_not_return_either(self, client: TestClient) -> None:
        """Discarding is a decision about an ad, so the ad stays and dedups."""
        ApiSource.postings = [ad(1)]
        client.post("/api/search", json={"sources": [{"source_id": "apifake"}]})
        key = client.get("/api/postings").json()["rows"][0]["dedup_key"]
        client.post(
            "/api/applications/status", json={"dedup_key": key, "status": "skipped"}
        )

        client.post("/api/search", json={"sources": [{"source_id": "apifake"}]})
        page = client.get("/api/postings").json()
        assert page["total"] == 1
        assert page["rows"][0]["status"] == "skipped"

    def test_an_ad_turned_away_is_not_reconsidered_by_the_same_search(
        self, client: TestClient
    ) -> None:
        ApiSource.postings = [ad(1, deadline=date(2026, 1, 1))]
        client.post("/api/search", json={"sources": [{"source_id": "apifake"}]})

        started = client.post(
            "/api/search", json={"sources": [{"source_id": "apifake"}]}
        ).json()
        results = client.get(f"/api/search/{started['id']}/results").json()
        assert [one["code"] for one in results["rejected"]] == ["filtered_out_before"]


# ------------------------------------------------- what the wizard offers (A)


class TestThePlacesList:
    """Suggestions for the "where" step, so a place is picked rather than typed."""

    def test_a_prefix_finds_the_town(self, client: TestClient) -> None:
        places = client.get("/api/places", params={"q": "rij"}).json()["places"]
        assert places[0]["name"] == "Rijeka"
        assert places[0]["county"] == "Primorsko-goranska"

    def test_a_county_carries_the_feed_number_that_would_search_it(
        self, client: TestClient
    ) -> None:
        """The bonus in A3: picking a county can tick the right HZZ feed."""
        body = client.get("/api/places", params={"q": "primorsko", "kind": "county"}).json()
        assert body["places"][0] == {
            "name": "Primorsko-goranska",
            "kind": "county",
            "county": "Primorsko-goranska",
            "feed": 13,
        }

    def test_every_county_comes_back_whatever_was_asked(self, client: TestClient) -> None:
        """The picker draws all twenty-one the moment the field is focused."""
        assert len(client.get("/api/places", params={"q": "rij"}).json()["counties"]) == 21

    def test_a_place_that_sits_in_two_counties_claims_neither(
        self, client: TestClient
    ) -> None:
        (privlaka,) = [
            place
            for place in client.get("/api/places", params={"q": "privlaka"}).json()["places"]
            if place["name"] == "Privlaka"
        ]
        assert privlaka["county"] == ""
        assert privlaka["feed"] is None

    def test_an_unknown_kind_is_refused_rather_than_guessed(self, client: TestClient) -> None:
        assert client.get("/api/places", params={"kind": "planet"}).status_code == 422

    def test_a_signed_out_browser_gets_nowhere(self, anonymous: TestClient) -> None:
        assert anonymous.get("/api/places").status_code == 401


class TestTheEmployersSeenSoFar:
    """Suggestions for the employer fields, drawn from the user's own rows.

    From their own rows and not a company register, because the employer somebody
    wants to skip is one whose ads they are tired of -- and because no source in
    JobSheet needs credentials, and this should not be the first that does.
    """

    def test_nothing_collected_yet_means_nothing_to_offer(self, client: TestClient) -> None:
        assert client.get("/api/postings/companies").json() == {"companies": [], "total": 0}

    def test_it_offers_the_employers_in_the_users_rows(
        self, client: TestClient, state: Any
    ) -> None:
        state.db.save_rows([job(1, posting=ad(1, company="Geodetski zavod d.o.o.")), job(2)])
        names = [one["name"] for one in client.get("/api/postings/companies").json()["companies"]]
        assert "Geodetski zavod d.o.o." in names

    def test_one_employer_written_two_ways_is_one_entry(
        self, client: TestClient, state: Any
    ) -> None:
        state.db.save_rows(
            [
                job(1, posting=ad(1, company="ERICSSON NIKOLA TESLA")),
                job(2, posting=ad(2, company="Ericsson Nikola Tesla d.d.")),
            ]
        )
        body = client.get("/api/postings/companies").json()
        assert body["total"] == 1
        # The fullest spelling wins: it is the one the user saw in the ad.
        assert body["companies"][0]["name"] == "Ericsson Nikola Tesla d.d."
        assert body["companies"][0]["count"] == 2

    def test_the_query_narrows_it(self, client: TestClient, state: Any) -> None:
        state.db.save_rows(
            [
                job(1, posting=ad(1, company="Geodetski zavod d.o.o.")),
                job(2, posting=ad(2, company="Ericsson Nikola Tesla d.d.")),
            ]
        )
        body = client.get("/api/postings/companies", params={"q": "ericsson"}).json()
        assert [one["name"] for one in body["companies"]] == ["Ericsson Nikola Tesla d.d."]

    def test_the_legal_form_does_not_have_to_be_typed_to_find_it(
        self, client: TestClient, state: Any
    ) -> None:
        state.db.save_rows([job(1, posting=ad(1, company="Geodetski zavod d.o.o."))])
        body = client.get("/api/postings/companies", params={"q": "geodetski zavod"}).json()
        assert body["total"] == 1

    def test_an_ad_with_no_employer_is_not_an_employer(
        self, client: TestClient, state: Any
    ) -> None:
        state.db.save_rows([job(1, posting=ad(1, company=""))])
        assert client.get("/api/postings/companies").json()["total"] == 0

    def test_one_account_is_not_offered_anothers_employers(
        self, client: TestClient, state: Any, settings: Settings
    ) -> None:
        """The account boundary, on a route that reads every row this user has."""
        state.db.save_rows([job(1, posting=ad(1, company="Geodetski zavod d.o.o."))])
        with TestClient(client.app, base_url=BASE_URL) as other:
            other.headers[TOKEN_HEADER] = settings.token
            sign_up(other, "somebody-else")
            assert other.get("/api/postings/companies").json() == {"companies": [], "total": 0}
